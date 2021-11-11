from fpdf import FPDF
from tkinter import *
from tkinter import ttk
import os
from tkinter import messagebox
import tkinter
from tkscrolledframe import ScrolledFrame
import pandas as pd
from datetime import date, datetime
import glob
from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileWriter, PdfFileReader
import win32com.client as win32
from openpyxl.reader.excel import load_workbook
import pickle

class Analise:

    def __init__(self, janela):
        self.janela = janela
        titulo = ' '
        self.janela.title(110 * titulo + 'Consultas')
        self.janela.geometry('800x500+350+100')
        self.janela.resizable(width=False, height=False)

        j = 0
        r = 0
        for i in range(800):
            c = str(222222 + r)
            Frame(self.janela, width=10, height=500, bg='#' + c).place(x=j, y=0)
            j += 10
            r += 1

        self.frame1 = Frame(self.janela, width=700, height=400, bd=7, bg='white', relief=RIDGE)
        self.frame1.place(x=50, y=50)

        lblini = Label(self.frame1, text='ANÁLISE TRIBUTÁRIA', font=('arial', 26, 'bold'), bd=0, bg='white'). \
            place(x=155, y=30)
        lblini = Label(self.frame1, text='Selecione a opção:', font=('arial', 14, 'bold'), bd=0, bg='white').\
            place(x=235, y=120)

        btn1 = Button(self.frame1, text='Análises \nPendentes', font=('arial', 16, 'bold'), bg='#FF5733',
                      fg='white', bd=4, width=12, justify=CENTER, command=self.pendentes).place(x=30, y=180)
        btn2 = Button(self.frame1, text='Nova \nAnálise', font=('arial', 16, 'bold'), width=12, bg='#FF5733', bd=4,
                      fg='white', command= lambda:[self.tela_principal(), self.tela_servicos(),
                    self.servicos.withdraw(), self.tela_materiais(), self.materiais.withdraw(),
                    self.tela_observacoes(), self.observacoes.withdraw(), self.tela_contratos(),
                                                   self.contratos.withdraw()]).place(x=250, y=180)

        btn3 = Button(self.frame1, text='Carregar \nAnálise', font=('arial', 16, 'bold'), bg='#FF5733',
                      fg='white', bd=4, width=12, justify=CENTER, command=lambda:[self.carregar_analise(),
                    self.tela_principal(), self.principal.withdraw(), self.tela_servicos(), self.servicos.withdraw(), self.tela_materiais(),
                    self.materiais.withdraw(), self.tela_observacoes(), self.observacoes.withdraw(), self.tela_contratos(),
                    self.contratos.withdraw()]).place(x=470, y=180)


    def pendentes(self):
        self.pendente = Toplevel()
        titulo = ' '
        self.pendente.title(100 * titulo + 'Consultas')
        self.pendente.geometry('800x500+350+100')
        self.pendente.resizable(width=False, height=False)

        estilo = ttk.Style()
        estilo.theme_use('default')
        estilo.configure('Treeview', background='#D3D3D3', foreground='black', rowheight=25,
                         fieldbackground='#D3D3D3')
        estilo.map('Treeview', background=[('selected', '#347083')])

        # Treeview frame
        tree_frame = Frame(self.pendente)
        tree_frame.pack(pady=100)
        # Barra rolagem
        tree_scroll = Scrollbar(tree_frame)
        tree_scroll.pack(side=RIGHT, fill=Y)
        # Criar Treeview
        nf_tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, selectmode='extended')
        nf_tree.pack(side=LEFT)
        # Configurar Barra Rolagem
        tree_scroll.config(command=nf_tree.yview)
        # Definir colunas
        colunas2 = ['Análise', 'Data']
        nf_tree['columns'] = colunas2
        # formatar colunas
        nf_tree.column('Análise', width=350)
        nf_tree.column('Data', width=100)

        # formatar títulos
        nf_tree.heading('Análise', text='Análise', anchor=W)
        nf_tree.heading('Data', text='Data', anchor=W)

        nf_tree['show'] = 'headings'

        self.lista = []
        # lista = [['DV-004-2021', '29/09/2021'], ['IN-006-2021', '30/09/2021']]
        self.pasta1 = os.listdir('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes')
        self.pasta = []
        for item in self.pasta1:
            if item.endswith('.pdf') is True and item != 'watermark.pdf':
                self.pasta.append(item)



        for i, n in enumerate(self.pasta):
            mod = os.path.getctime('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes')
            mod = datetime.fromtimestamp(mod)
            data = date.strftime(mod, '%d/%m/%Y')
            self.lista.append([])
            self.lista[i].append(n)
            self.lista[i].append(data)

        # print(self.lista)
        # inserir dados do banco no treeview
        def inserir_tree(lista):
            nf_tree.delete(*nf_tree.get_children())
            contagem = 0
            for row in lista:  # loop para inserir cores diferentes nas linhas
                if contagem % 2 == 0:
                    nf_tree.insert(parent='', index='end', text='', iid=contagem,
                                   values=(row[0], row[1]), tags=('evenrow',))
                else:
                    nf_tree.insert(parent='', index='end', text='', iid=contagem,
                                   values=(row[0], row[1]), tags=('oddrow',))
                contagem += 1


            py = 115
            self.list_check = []
            for lin in self.lista:
                self.check1 = IntVar()
                lbl1 = Checkbutton(self.pendente, var=self.check1, onvalue=1, offvalue=0)
                lbl1.place(y=py, x=138)
                py += 26
                self.list_check.append(self.check1)



        # Create the watermark from an image
        c = canvas.Canvas('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\watermark.pdf')
        # Draw the image at x, y. I positioned the x,y to be where i like here
        c.drawImage('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Paulo.png', 440, 30, 100, 60, mask='auto')
        c.save()
        # Get the watermark file you just created
        watermark = PdfFileReader(open("G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\watermark.pdf", "rb"))
        # Get our files ready



        def assinatura():
            caminho = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
            self.salvos = []
            for n, arquivo in enumerate(self.pasta):
                if self.list_check[n].get() == 1:
                    dir_atual = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
                    os.chdir(dir_atual)

                    # if n == 0:
                    #     # Create the watermark from an image
                    #     c = canvas.Canvas('watermark.pdf')
                    #     # Draw the image at x, y. I positioned the x,y to be where i like here
                    #     c.drawImage('Paulo.png', 440, 30, 100, 60, mask='auto')
                    #     c.save()
                    #     # Get the watermark file you just created
                    #     watermark = PdfFileReader(open("watermark.pdf", "rb"))
                    #     # Get our files ready
                    #     output_file = PdfFileWriter()
                    self.output_file = PdfFileWriter()
                    with open(caminho + arquivo, "rb") as f:
                        input_file = PdfFileReader(f, "rb")
                        # Number of pages in input document
                        page_count = input_file.getNumPages()

                        # Go through all the input file pages to add a watermark to them
                        for page_number in range(page_count):
                            input_page = input_file.getPage(page_number)
                            if page_number == page_count - 1:
                                input_page.mergePage(watermark.getPage(0))
                            self.output_file.addPage(input_page)

                        # dir = os.getcwd()
                        path = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021'
                        os.chdir(path)
                        file = glob.glob(str(arquivo[21:32]) + '*')
                        file = ''.join(file)
                        try:
                            os.chdir(file)
                        except:
                            os.chdir(path)

                        # finally, write "output" to document-output.pdf
                        with open('Análise Tributária - ' + str(arquivo[21:]), "wb") as outputStream:
                            self.output_file.write(outputStream)

                    os.chdir(caminho)
                    os.remove(arquivo)
                    self.salvos.append(n)


            troca = 0
            for i in self.salvos:
                self.pasta.pop(i-troca)
                self.lista.pop(i-troca)
                troca += 1


            outlook = win32.Dispatch('outlook.application')

            # criar um email
            email = outlook.CreateItem(0)

            # configurar as informações do seu e-mail
            email.To = "loliveira@gasbrasiliano.com.br"
            email.Subject = "E-mail automático Análise Tributária"
            email.HTMLBody = f"""
                        <p>Análise(s) Tributária(s) assinada(s) com sucesso.</p>

                        """

            email.Send()
            tkinter.messagebox.showinfo('', 'Análise(s) assinada(s) com sucesso!')
            # self.pendente.update()
            self.pendente.lift()


            self.list_check.clear()
            inserir_tree(self.lista)



        def recusa():
            resultado = tkinter.messagebox.askquestion('', 'Deseja Justificar?')
            if resultado == 'yes':
                self.justifica = Toplevel()
                self.justifica.geometry('500x300+500+200')
                Label(self.justifica, text='Justificativa:', font=('arial', 14, 'bold')).grid(row=0, column=0, padx=35,
                                                                                              pady=10, sticky=W)
                self.texto_just = Text(self.justifica, width=50, height=10, wrap=WORD)
                self.texto_just.grid(row=1, column=0, padx=40)
                def email():
                    outlook = win32.Dispatch('outlook.application')

                    # criar um email
                    email = outlook.CreateItem(0)

                    # configurar as informações do seu e-mail
                    email.To = "loliveira@gasbrasiliano.com.br"
                    email.Subject = "E-mail automático Análise Tributária"
                    email.HTMLBody = f"""
                                            <p>Análise Tributária foi recusada.</p>

                                            """

                    email.Send()
                    tkinter.messagebox.showinfo('', 'Análise(s) assinada(s) com sucesso!')
                    self.pendente.lift()

                envio = Button(self.justifica, text='Enviar', font=('arial', 12, 'bold'), bd=2).grid(row=2, column=0, pady=20)

            else:
                tkinter.messagebox.showinfo('', 'Recusado com Sucesso!')

        assinar = Button(self.pendente, text='Assinar', font=('arial', 14, 'bold'), width=10, bd=3, command=assinatura).\
            place(x=250, y=400)

        recusar = Button(self.pendente, text='Recusar', font=('arial', 14, 'bold'), width=10, bd=3, command=recusa). \
            place(x=420, y=400)

        def NotasInfo2(ev):
            # fn_id.delete(0, END)
            verinfo2 = nf_tree.focus()
            dados2 = nf_tree.item(verinfo2)
            row = dados2['values']
            print(row)
            # entr_atual.delete(0, END)
            # entr_atual.insert(0, row[3])
            # fn_id.insert(0, row[0])
            os.startfile('\\\GBD_VT1NTAQA\Data2\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes' + '\\' + row[0])

        # adicionar a tela
        nf_tree.tag_configure('oddrow', background='white')
        nf_tree.tag_configure('evenrow', background='lightblue')
        inserir_tree(self.lista)
        nf_tree.bind('<Double-Button>', NotasInfo2)


    def carregar_analise(self):
        self.carregar = Toplevel()
        titulo = ' '
        self.carregar.title(160 * titulo + 'Análise Tributária')
        self.carregar.geometry('1100x680+200+20')
        self.carregar.resizable(width=False, height=False)

        estilo = ttk.Style()
        estilo.theme_use('default')
        estilo.configure('Treeview', background='#D3D3D3', foreground='black', rowheight=25,
                         fieldbackground='#D3D3D3')
        estilo.map('Treeview', background=[('selected', '#347083')])

        # Treeview frame
        tree_frame1 = Frame(self.carregar)
        tree_frame1.pack(pady=100)
        # Barra rolagem
        tree_scroll1 = Scrollbar(tree_frame1)
        tree_scroll1.pack(side=RIGHT, fill=Y)
        # Criar Treeview
        nf_tree1 = ttk.Treeview(tree_frame1, yscrollcommand=tree_scroll1.set, selectmode='extended')
        nf_tree1.pack(side=LEFT)
        # Configurar Barra Rolagem
        tree_scroll1.config(command=nf_tree1.yview)
        # Definir colunas
        colunas2 = ['Análise', 'Data']
        nf_tree1['columns'] = colunas2
        # formatar colunas
        nf_tree1.column('Análise', width=200)
        nf_tree1.column('Data', width=200)

        # formatar títulos
        estilo.configure("Treeview.Heading", font=('arial', 11), background='DodgerBlue3', foreground='white')
        nf_tree1.heading('Análise', text='Análise', anchor=W)
        nf_tree1.heading('Data', text='Data', anchor=W)

        nf_tree1['show'] = 'headings'
        lista = []
        temp_list = []
        with open('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Base.txt', "rb") as carga:
            while True:
                try:
                    temp_list.append(pickle.load(carga))
                except EOFError:
                    break
        for n, item in enumerate(temp_list):
            lista.append([])
            lista[n].append(item[2])
            lista[n].append(item[0])


        # inserir dados do banco no treeview
        def inserir_tree(lista):
            nf_tree1.delete(*nf_tree1.get_children())
            contagem = 0
            lista.sort(key=lambda lista: datetime.strptime(lista[1], '%d/%m/%Y, %H:%M:%S'), reverse=True)
            print(lista)
            for row in lista:  # loop para inserir cores diferentes nas linhas
                if contagem % 2 == 0:
                    nf_tree1.insert(parent='', index='end', text='', iid=contagem,
                                    values=(row[0], row[1]), tags=('evenrow',))
                else:
                    nf_tree1.insert(parent='', index='end', text='', iid=contagem,
                                    values=(row[0], row[1]), tags=('oddrow',))
                contagem += 1





        def NotasInfo2(ev):
            # fn_id.delete(0, END)
            verinfo3 = nf_tree1.focus()
            dados3 = nf_tree1.item(verinfo3)
            temp_list.sort(key=lambda lista: datetime.strptime(lista[0], '%d/%m/%Y, %H:%M:%S'), reverse=True)
            self.gere.delete(0, END)
            self.proc.delete(0, END)
            self.req.delete(0, END)
            self.orcam.delete(0, END)
            self.objcust.delete(0, END)
            self.tipoa.deselect()
            self.tipob.deselect()
            self.tipoc.deselect()
            self.objeto.delete(1.0, END)
            self.valor.delete(0, END)
            self.complem.delete(1.0, END)
            self.linha_mat.delete(0, END)
            self.cod_serv.delete(0, END)
            self.iva.delete(0, END)
            self.linha_serv.delete(0, END)
            self.obs.delete(1.0, END)
            self.obs_serv.delete(1.0, END)
            self.obs1.delete(1.0, END)
            self.obs2.delete(1.0, END)
            [i.delete(1.0, END) for i in self.infos]
            # [i.deselect() for i in self.var_check]
            self.gere.insert(0, temp_list[int(verinfo3)][1])
            self.proc.insert(0, temp_list[int(verinfo3)][2])
            self.req.insert(0, temp_list[int(verinfo3)][3])
            self.orcam.insert(0, temp_list[int(verinfo3)][4])
            self.objcust.insert(0, temp_list[int(verinfo3)][5])
            self.tipo1.set(int(temp_list[int(verinfo3)][6]))
            self.tipo2.set(int(temp_list[int(verinfo3)][7]))
            self.tipo3.set(int(temp_list[int(verinfo3)][8]))
            self.objeto.insert(1.0, temp_list[int(verinfo3)][9].strip())
            self.valor.insert(0, temp_list[int(verinfo3)][10])
            self.complem.insert(1.0, temp_list[int(verinfo3)][11].strip())
            for r, val in enumerate(temp_list[int(verinfo3)][12][1:]):
                print(val)
                for b, value in enumerate(val):
                    print(value)
                    self.lista_mat[b][r].delete(0, END)
                    self.lista_mat[b][r].insert(0, value)

            self.linha_mat.insert(0, temp_list[int(verinfo3)][13])
            self.serv.insert(1.0, temp_list[int(verinfo3)][14].strip())
            self.iva.insert(0, temp_list[int(verinfo3)][15])
            cont = 0
            for r, val in enumerate(temp_list[int(verinfo3)][16][1:]):
                for b, value in enumerate(val):
                    if cont == 0:
                        self.lista[b+1][r].delete(0, END)
                        self.lista[b+1][r].insert(0, value)
                    elif cont == 1:
                        self.lista[b-1][r].delete(0, END)
                        self.lista[b-1][r].insert(0, value)
                    else:
                        self.lista[b][r].delete(0, END)
                        self.lista[b][r].insert(0, value)
                    cont += 1
                cont = 0
            self.linha_serv.insert(0, temp_list[int(verinfo3)][17])
            self.obs.insert(1.0, temp_list[int(verinfo3)][18].strip())
            self.obs_serv.insert(1.0, temp_list[int(verinfo3)][19])
            self.obs1.insert(1.0, temp_list[int(verinfo3)][20].strip())
            self.obs2.insert(1.0, temp_list[int(verinfo3)][21].strip())
            for n, i in enumerate(temp_list[int(verinfo3)][22]):
                self.infos[n].insert(1.0, i)
                print(i)
            for n, i in enumerate(temp_list[int(verinfo3)][23]):
                if i == 1:
                    self.lista_check[n].set(1)


            self.carregar.destroy()
            self.principal.deiconify()

        # adicionar a tela
        nf_tree1.tag_configure('oddrow', background='white')
        nf_tree1.tag_configure('evenrow', background='lightblue')
        inserir_tree(lista)
        nf_tree1.bind('<Double-Button>', NotasInfo2)




    def tela_principal(self):
        self.principal = Toplevel()
        titulo = ' '
        self.principal.title(160 * titulo + 'Análise Tributária')
        self.principal.geometry('1100x680+200+20')
        self.principal.resizable(width=False, height=False)

        self.mainframe = Frame(self.principal, width=1200, height=680, relief=RIDGE, bg='DeepSkyBlue3')
        self.mainframe.place(x=0, y=0)

        info_frame = Frame(self.mainframe, width=1080, height=260, relief=RIDGE, bd=7)
        info_frame.place(x=10, y=30)
        info_frame2 = Frame(self.mainframe, width=1080, height=300, relief=RIDGE, bd=7)
        info_frame2.place(x=10, y=290)
        info_frame3 = Frame(self.mainframe, width=1080, height=70, relief=RIDGE, bd=5)
        info_frame3.place(x=10, y=590)
        self.fonte = ('arial', 12, 'bold')
        self.ent_fonte = ('arial', 12)

        self.rotulos = ['Gerência Contratante: ', 'Nº Processo GECBS: ', 'Requisição de Compras: ', 'Código Material/Serviço:',
                   'Consta no Orçamento?', 'Sim', 'Não', '1.	Classificação Contábil:',
                   '1.	Objeto de Custos:', 'Serviço', 'Material ', 'Serviço com Fornecimento de Material']

        Label(info_frame, text='Gerência Contratante: ', font=self.fonte, bd=0).place(x=20, y=50)
        Label(info_frame, text='Nº Processo GECBS: ', font=self.fonte, bd=0).place(x=500, y=50)
        Label(info_frame, text='Requisição de Compras: ', font=self.fonte, bd=0).place(x=20, y=90)
        Label(info_frame, text='Consta no Orçamento?', font=self.fonte, bd=0).place(x=500, y=90)
        Label(info_frame, text='Objeto de Custos:', font=self.fonte, bd=0).place(x=20, y=130)
        Label(info_frame, text='Tipo de Análise: ', font=self.fonte, bd=0).place(x=500, y=130)
        Label(info_frame2, text='Objeto da Contratação: ', font=self.fonte, bd=0).place(x=20, y=20)
        Label(info_frame2, text='Valor estimado: ', font=self.fonte, bd=0).place(x=550, y=20)
        Label(info_frame2, text='Informações Complementares: ', font=self.fonte, bd=0).place(x=20, y=175)

        self.gere = Entry(info_frame, width=20, bd=4, font=self.ent_fonte)
        self.gere.place(x=230, y=48)
        self.proc = Entry(info_frame, width=20, bd=4, font=self.ent_fonte)
        self.proc.place(x=700, y=50)
        self.req = Entry(info_frame, width=20, bd=4, font=self.ent_fonte)
        self.req.place(x=230, y=88)
        self.orcam = ttk.Combobox(info_frame, font=self.fonte, width=12)
        self.orcam['values'] = ('Não', 'Sim')
        self.orcam.current(1)
        self.orcam.place(x=700, y=88)
        self.objcust = Entry(info_frame, width=20, bd=4, font=self.ent_fonte)
        self.objcust.place(x=230, y=128)
        self.tipo1 = IntVar()
        self.tipo2 = IntVar()
        self.tipo3 = IntVar()
        self.tipoa = Checkbutton(info_frame, var=self.tipo1, onvalue=1, offvalue=0, text='Serviço', font=('arial', 12))
        self.tipoa.place(x=690, y=125)
        self.tipob = Checkbutton(info_frame, var=self.tipo2, onvalue=1, offvalue=0, text='Material', font=('arial', 12))
        self.tipob.place(x=690, y=155)
        self.tipoc = Checkbutton(info_frame, var=self.tipo3, onvalue=1, offvalue=0, text='Serviço com Fornc. Material',
                                 font=('arial', 12))
        self.tipoc.place(x=690, y=185)
        self.objeto = Text(info_frame2, width=50, height=5, bd=4, font='arial')
        self.objeto.place(x=20, y=55)
        self.valor = Entry(info_frame2, width=20, bd=4, font=self.ent_fonte)
        self.valor.place(x=550, y=55)
        self.complem = Text(info_frame2, width=50, height=3, bd=4, font='arial')
        self.complem.place(x=20, y=200)


        self.chama_mat = Button(info_frame3, font=('arial', 14, 'bold'), text='Materiais', bd=3, width=20,
                                     command=lambda :[self.principal.withdraw(), self.materiais.deiconify()])
        self.chama_mat.place(x=20, y=2)


        self.chama_serv = Button(info_frame3, font=('arial', 14, 'bold'), text='Serviços', bd=3, width=20,
                                command=lambda :[self.principal.withdraw(), self.servicos.deiconify()])
        self.chama_serv.place(x=270, y=2)

        self.chama_obs2 = Button(info_frame3, font=('arial', 14, 'bold'), text='Observações', bd=3, width=20,
                        command=lambda :[self.principal.withdraw(),self.observacoes.deiconify()]).place(x=520, y=2)

        self.chama_contr = Button(info_frame3, font=('arial', 14, 'bold'), text='Clausulas', bd=3, width=20,
                        command=lambda :[self.principal.withdraw(), self.contratos.deiconify()]).place(x=770, y=2)




    def tela_servicos(self):
        self.servicos = Toplevel()
        titulo = ' '
        self.servicos.title(160 * titulo + 'Serviços')
        self.servicos.geometry('1100x690+200+20')
        self.servicos.config(bg='DeepSkyBlue2')

        self.serv_frame = Frame(self.servicos, width=1080, height=600, relief=RIDGE, bd=7)
        self.serv_frame.place(x=10, y=10)
        self.serv_frame1 = Frame(self.servicos, width=1080, height=70, relief=RIDGE, bd=7)
        self.serv_frame1.place(x=10, y=610)

        # frame_1 = Frame(serv_frame, height=500, width=1190, bd=5).place(x=0, y=0)
        Label(self.serv_frame, text='Codigo Serviço', font=self.fonte, bd=0).place(x=20, y=30)
        Label(self.serv_frame, text='Codigo IVA', font=self.fonte, bd=0).place(x=20, y=190)


        self.cod_serv = Entry(self.serv_frame, width=7, font=self.fonte, bd=3)
        self.cod_serv.place(x=150, y=30)
        self.serv = Text(self.serv_frame, width=70, height=7, bd=4, font='arial', wrap=WORD)
        self.serv.place(x=240, y=30)

        self.iva = Entry(self.serv_frame, width=10, bd=4, font=self.fonte)
        self.iva.place(x=150, y=190)
        Label(self.serv_frame, text='Quebra', font=('arial', 10, 'bold'), bd=0).place(x=990, y=505)
        self.linha_serv = Entry(self.serv_frame, font=('arial', 10, 'bold'), bd=2, width=3)
        self.linha_serv.place(x=1000, y=530)
        self.linha_serv.insert(0, 0)

        def busca_servico(ev):
            self.path = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
            data_serv = pd.read_excel(self.path + 'material.xlsx', sheet_name='116')
            data_serv = pd.DataFrame(data_serv)
            for index, row in data_serv.iterrows():
                if self.cod_serv.get() == row['servico']:
                    self.serv.delete(1.0, END)
                    self.serv.insert(INSERT, row['servico'] + ' - ' + data_serv.loc[index, 'descricao'] + '\n')
                    self.serv.insert(INSERT, '\n')
                    self.serv.insert(INSERT, data_serv.loc[index, 'irrf'] + '\n')
                    self.serv.insert(INSERT, data_serv.loc[index, 'crf'] + '\n')
                    self.serv.insert(INSERT, data_serv.loc[index, 'inss'] + '\n')
                    self.serv.insert(INSERT, data_serv.loc[index, 'iss'] + '\n')



        self.cod_serv.bind('<FocusOut>', busca_servico)

        sf = ScrolledFrame(self.servicos, width=570, height=250)
        sf.place(x=285, y=270)

        # Bind the arrow keys and scroll wheel
        sf.bind_arrow_keys(self.servicos)
        sf.bind_scroll_wheel(self.servicos)

        # Create a frame within the ScrolledFrame
        inner_frame2 = sf.display_widget(Frame)


        Label(self.serv_frame, text='Código', font=self.fonte).place(x=305, y=230)
        Label(self.serv_frame, text='Descrição', font=self.fonte).place(x=505, y=230)
        Label(self.serv_frame, text='C.C.', font=self.fonte).place(x=750, y=230)
        self.lista = [[], [], []]
        self.entradas = []
        self.data = [['DESCRIÇÃO', 'CÓDIGO', 'C.C']]
        # self.data.append(['MONITORAMENTO ALARME - OP', '3001580', '5605'])
        # múltiplos serviços
        px = 265
        py = 255
        for i in range(60):
            for c in range(3):
                if c == 1:
                    largura = 30
                    espaco = 280
                else:
                    largura = 15
                    espaco = 145
                serv = Entry(inner_frame2, width=largura, bd=3, font='arial')
                serv.grid(row=i, column=c)
                self.entradas.append(serv)
                self.lista[c].append(serv)
                px += espaco
            py += 26
            px = 265



        def colar(ev):
            if self.servicos.winfo_viewable() == 1:
                path = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
                serv_cad = pd.read_excel(path + 'material.xlsx', sheet_name='servicos')
                serv_cad = pd.DataFrame(serv_cad)
                serv_cad['Nº de serviço'] = serv_cad['Nº de serviço'].astype(str)
                rows = self.servicos.clipboard_get().split('\n')
                self.entradas[0].clipboard_clear()

                rows.pop() if len(rows) > 1 else rows

                for r, val in enumerate(rows):
                    values = val.split('\t')
                    # ordem = [1, 0, 2]
                    # values = [values[i] for i in ordem]
                    if len(values) > 1:
                        del values[1:]
                    for b, value in enumerate(values):
                        for index, row in serv_cad.iterrows():
                            self.lista[b][r].delete(0, END)
                            self.lista[b][r].insert(0, value)
                            if value == row['Nº de serviço']:
                                self.lista[b + 1][r].insert(0, serv_cad.loc[index, 'Denominação'])
                                self.lista[b + 2][r].insert(0, int(serv_cad.loc[index, 'Classe avaliaç.']))
            else:
                pass




        def adicionar():


            cont2 = 0
            temp_list = []
            for lin in self.entradas:
                if lin.get() != '':
                    temp_list.append(lin.get())

                    cont2 += 1
                    if cont2 == 3:
                        ordem = [1, 0, 2]
                        temp_list = [temp_list[i] for i in ordem]
                        lista_nova = temp_list.copy()
                        self.data.append(lista_nova)
                        temp_list.clear()
                        cont2 = 0

            for lin in self.entradas:
                lin.delete(0, END)







        def limpar():
            del self.data[1:]
            for lin in self.entradas:
                lin.delete(0, END)


        self.entradas[0].bind("<<Paste>>", colar)





        # epw = pdf.w - 2 * pdf.l_margin
        # col_width = epw / 3
        # data2 = ['CÓDIGO', 'DESCRIÇÃO', 'C.C']


        self.btnadicionar = Button(self.serv_frame, font=self.fonte, text='Adicionar', bd=4, width=15,
                                   command=adicionar).place(x=380, y=530)

        self.btnlimpar = Button(self.serv_frame, font=self.fonte, text='Limpar campos', bd=4, width=15,
                                command=limpar).place(x=570, y=530)

        self.chama_princ1 = Button(self.serv_frame1, font=('arial', 14, 'bold'), text='Dados Iniciais', bd=3, width=20,
                                  command=lambda: [self.principal.deiconify(), self.servicos.withdraw()]).place(x=20,
                                                                                                                 y=2)

        self.chama_mat = Button(self.serv_frame1, font=('arial', 14, 'bold'), text='Materiais', bd=3, width=20,
                                  command=lambda: [self.materiais.deiconify(), self.servicos.withdraw(),
                                                   self.principal.withdraw()]).place(x=270, y=2)

        self.chama_obs1 = Button(self.serv_frame1, font=('arial', 14, 'bold'), text='Observações', bd=3, width=20,
                                command=lambda: [self.observacoes.deiconify(), self.servicos.withdraw(),
                                                 self.principal.withdraw()]).place(x=520, y=2)

        self.chama_contr3 = Button(self.serv_frame1, font=('arial', 14, 'bold'), text='Clausulas', bd=3, width=20,
                                   command=lambda: [self.contratos.deiconify(), self.servicos.withdraw(),
                                                    self.principal.withdraw()]).place(x=770, y=2)




    def tela_materiais(self):
        self.materiais = Toplevel()
        titulo = ' '
        self.materiais.title(160 * titulo + 'Materiais')
        self.materiais.geometry('1100x680+200+20')
        self.materiais.config(bg='DeepSkyBlue3')

        self.mat_frame = Frame(self.materiais, width=1080, height=580, relief=RIDGE, bd=7)
        self.mat_frame.place(x=10, y=10)
        self.mat_frame1 = Frame(self.materiais, width=1080, height=70, relief=RIDGE, bd=7)
        self.mat_frame1.place(x=10, y=585)

        # frame_1 = Frame(serv_frame, height=500, width=1190, bd=5).place(x=0, y=0)
        Label(self.mat_frame, text='Codigo', font=('arial', 10, 'bold'), bd=0).place(x=180, y=135)
        Label(self.mat_frame, text='Descrição', font=('arial', 10, 'bold'), bd=0).place(x=360, y=135)
        Label(self.mat_frame, text='IVA', font=('arial', 10, 'bold'), bd=0).place(x=545, y=135)
        Label(self.mat_frame, text='NCM', font=('arial', 10, 'bold'), bd=0).place(x=620, y=135)
        Label(self.mat_frame, text='ICMS', font=('arial', 10, 'bold'), bd=0).place(x=700, y=135)
        Label(self.mat_frame, text='IPI', font=('arial', 10, 'bold'), bd=0).place(x=770, y=135)
        Label(self.mat_frame, text='PIS', font=('arial', 10, 'bold'), bd=0).place(x=820, y=135)
        Label(self.mat_frame, text='COFINS', font=('arial', 10, 'bold'), bd=0).place(x=860, y=135)
        Label(self.mat_frame, text='Informações Tributárias', font=self.fonte, bd=0).place(x=150, y=50)
        Label(self.mat_frame, text='Quebra', font=('arial', 10, 'bold'), bd=0).place(x=990, y=505)
        self.linha_mat = Entry(self.mat_frame, font=('arial', 10, 'bold'), bd=2, width=3)
        self.linha_mat.place(x=1000, y=530)
        self.linha_mat.insert(0, 0)

        sf = ScrolledFrame(self.materiais, width=780, height=300)
        sf.place(x=168, y=175)

        # Bind the arrow keys and scroll wheel
        sf.bind_arrow_keys(self.materiais)
        sf.bind_scroll_wheel(self.materiais)

        # Create a frame within the ScrolledFrame
        inner_frame1 = sf.display_widget(Frame)



        self.lista_mat = [[], [], [], [], [], [], [], []]
        self.entradas_mat = []
        self.data_mat = [['CÓDIGO', 'DESCRIÇÃO', 'IVA', 'NCM', 'ICMS', 'IPI', 'PIS', 'COFINS']]
        self.mat_check = []
        # listin = ['604400','CENTRAL DE ALARME COMPLETA COM 18','ZB','8531.10.10','18,0%','15,0%','1,65%','7,6%']
        # for l in range(22):
        #     self.data_mat.append(listin)


        # múltiplos materiais
        px = 50
        py = 160
        for i in range(30):
            for c in range(8):
                if c == 1:
                    largura = 30
                    espaco = 280
                elif c == 0:
                    largura = 10
                    espaco = 100
                elif c == 3:
                    largura = 10
                    espaco = 100
                else:
                    largura = 5
                    espaco = 55
                mater = Entry(inner_frame1, width=largura, bd=4, font='arial')
                mater.grid(row=i, column=c)
                self.entradas_mat.append(mater)
                self.lista_mat[c].append(mater)
                px += espaco
            py += 30
            px = 50

        self.check2 = IntVar()
        self.check3 = IntVar()
        self.check4 = IntVar()
        self.check2 = Checkbutton(self.mat_frame, variable=self.check2, onvalue=1,
                             offvalue=0, bd=0, font=('arial', 14))
        self.check2.place(x=545, y=100)
        self.check3 = Checkbutton(self.mat_frame, variable=self.check3, onvalue=1,
                                  offvalue=0, bd=0, font=('arial', 14))
        self.check3.place(x=625, y=100)
        self.check4 = Checkbutton(self.mat_frame, variable=self.check4, onvalue=1,
                                  offvalue=0, bd=0, font=('arial', 14))
        self.check4.place(x=780, y=100)


        # self.multi
        def preenche_aliq(ev):
            for e, item in enumerate(self.entradas_mat):
                if e % 8 == 0:
                    if item.get() != '':
                        self.entradas_mat[e + 4].delete(0, END)
                        self.entradas_mat[e + 4].insert(0, '18%')
                        self.entradas_mat[e + 5].delete(0, END)
                        self.entradas_mat[e + 5].insert(0, '15%')
                        self.entradas_mat[e+6].delete(0, END)
                        self.entradas_mat[e+6].insert(0, '1,65%')
                        self.entradas_mat[e + 7].delete(0, END)
                        self.entradas_mat[e + 7].insert(0, '7,6%')


        self.check4.bind('<Button>', preenche_aliq)

        def preenche_iva(ev):
            for e, item in enumerate(self.entradas_mat):
                if e % 8 == 0 and e != 0:
                    if item.get() != '':
                        self.entradas_mat[e + 2].delete(0, END)
                        self.entradas_mat[e + 2].insert(0, self.entradas_mat[2].get())


        self.check2.bind('<Button>', preenche_iva)

        def preenche_ncm(ev):
            for e, item in enumerate(self.entradas_mat):
                if e % 8 == 0 and e != 0:
                    if item.get() != '':
                        self.entradas_mat[e + 3].delete(0, END)
                        self.entradas_mat[e + 3].insert(0, self.entradas_mat[3].get())


        self.check3.bind('<Button>', preenche_ncm)



        def colar(ev):
            path = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
            cad_mat = pd.read_excel(path + 'material.xlsx', sheet_name='materiais')
            cad_mat = pd.DataFrame(cad_mat)
            cad_mat['Material'] = cad_mat['Material'].astype(str)
            rows = self.entradas_mat[0].clipboard_get().split('\n')
            self.entradas_mat[0].clipboard_clear()
            rows.pop()
            for r, val in enumerate(rows):
                values = val.split('\t')
                if len(values) > 1:
                    del values[1:]
                for b, value in enumerate(values):
                    for index, row in cad_mat.iterrows():
                        self.lista_mat[b][r].delete(0, END)
                        self.lista_mat[b][r].insert(0, value)
                        if value == row['Material']:
                            campo = cad_mat.loc[index, 'Texto breve material']
                            campo = campo[:32]
                            self.lista_mat[b+1][r].insert(0, campo)
                            self.lista_mat[b + 3][r].insert(0, cad_mat.loc[index, 'Ncm'])


        def adicionar():
            cont2 = 0
            mat_list = []
            for lin in self.entradas_mat:
                if lin.get() != '':
                    mat_list.append(lin.get())
                    cont2 += 1
                    if cont2 == 8:
                        lista_nova_mat = mat_list.copy()
                        self.data_mat.append(lista_nova_mat)
                        mat_list.clear()
                        cont2 = 0
            for lin in self.entradas_mat:
                lin.delete(0, END)




        def limpar():
            del self.data_mat[1:]
            for lin in self.entradas_mat:
                lin.delete(0, END)


        self.entradas_mat[0].bind("<<Paste>>", colar)


        # epw = pdf.w - 2 * pdf.l_margin
        # col_width = epw / 3
        # data2 = ['CÓDIGO', 'DESCRIÇÃO', 'C.C']


        self.btnadicionar = Button(self.mat_frame, width=15, font=self.fonte, text='Adicionar', bd=4,
                                   command=adicionar).place(x=350, y=480)

        self.btnlimpar = Button(self.mat_frame, width=15, font=self.fonte, text='Limpar campos', bd=4,
                                command=limpar).place(x=550, y=480)

        self.chama_princ = Button(self.mat_frame1, font=('arial', 14, 'bold'), text='Dados Iniciais', bd=3, width=20,
                                 command=lambda: [self.principal.deiconify(), self.materiais.withdraw()]).place(x=20, y=2)

        self.chama_serv2 = Button(self.mat_frame1, font=('arial', 14, 'bold'), text='Serviços', bd=3, width=20,
                                 command=lambda: [self.servicos.deiconify(), self.materiais.withdraw(),
                                                  self.principal.withdraw()]).place(x=270, y=2)

        self.chama_obs = Button(self.mat_frame1, font=('arial', 14, 'bold'), text='Observações', bd=3, width=20,
                                  command=lambda: [self.observacoes.deiconify(), self.materiais.withdraw(),
                                                   self.principal.withdraw()]).place(x=520, y=2)

        self.chama_contr2 = Button(self.mat_frame1, font=('arial', 14, 'bold'), text='Clausulas', bd=3, width=20,
                                 command=lambda: [self.contratos.deiconify(), self.materiais.withdraw(),
                                                  self.principal.withdraw()]).place(x=770, y=2)


        # self.avancar = Button(self.mat_frame, font=self.fonte, text='Avançar', bd=4,
        #                       command=self.contratos).place(x=1100, y=600)

    def tela_contratos(self):
        self.contratos = Toplevel()
        titulo = ' '
        self.contratos.title(160 * titulo + 'Serviços')
        self.contratos.geometry('1200x680+100+20')

        # self.cont_frame = Frame(self.contratos, width=1200, height=680, relief=RIDGE, bd=7, bg='floral white')
        # self.cont_frame.pack(fill=BOTH, expand=1)

        sf = ScrolledFrame(self.contratos, width=940, height=480)
        sf.pack(side="top", expand=1, fill="both")

        # Bind the arrow keys and scroll wheel
        sf.bind_arrow_keys(self.contratos)
        sf.bind_scroll_wheel(self.contratos)

        # Create a frame within the ScrolledFrame
        inner_frame = sf.display_widget(Frame)

        # Add a bunch of widgets to fill some space
        self.nomes = ['N/A', 'Minuta', 'Redação', '2.3.7.', '2.3.7.1', '2.3.7.2', '2.3.7.3', '2.3.7.4', '6.8.2', '15.1',
                      '3.10.1', '3.9-10-11', 'Anexo 2']
        self.lista_check = []
        linha = 3
        for check in range(13):
            self.lista_check.append(IntVar())
            check1 = Checkbutton(inner_frame, text=self.nomes[check], variable=self.lista_check[check], onvalue=1,
                                 offvalue=0, bd=0, font=('arial', 14))
            check1.grid(row=linha, column=1, padx=70)
            linha += 2


        self.info_lbl = Label(inner_frame, bd=0, bg='floral white', font=('arial', 14),
                              text='Informações  contratuais: ').grid(row=2, column=2, pady=30)


        os.getcwd()
        path = 'G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
        with open(path + 'texto.txt', 'r', encoding='latin-1') as read_obj:
            csv_reader = read_obj.readlines()
            self.infos = []
            linha = 2
            for row in csv_reader:
                Label(inner_frame).grid(row=linha, column=2)
                self.campo = Text(inner_frame, height=5, width=100, wrap=WORD)
                self.campo.grid(row=linha + 1, column=2)
                self.campo.insert('end', row)
                self.infos.append(self.campo)
                linha += 2

        def enviar_email():
            outlook = win32.Dispatch('outlook.application')

            # criar um email
            email = outlook.CreateItem(0)

            # configurar as informações do seu e-mail
            email.To = "pfranca@gasbrasiliano.com.br"
            email.Subject = "E-mail automático Análise Tributária"
            email.HTMLBody = f"""
            <p>Análise Tributária {self.proc.get()} está disponível para assinatura.</p>

            """

            # anexo = "C://Users/joaop/Downloads/arquivo.xlsx"
            # email.Attachments.Add(anexo)

            email.Send()
            tkinter.messagebox.showinfo('', 'Email enviado com sucesso')
            self.contratos.lift()

        self.voltar = Button(inner_frame, font=self.fonte, text='Voltar', bd=4,
                             command=lambda:[self.principal.deiconify(), self.contratos.withdraw()]).\
            grid(row=28, column=1, pady=10, padx=20)

        self.gerar = Button(inner_frame, font=self.fonte, text='Gerar PDF', bd=4,
                              command=self.salvar).grid(row=28, column=2, pady=30)

        self.enviar = Button(inner_frame, font=self.fonte, text='Enviar Email', bd=4,
                               command=enviar_email).grid(row=29, column=2, pady=10)

        Label(inner_frame, text='Quebra', font=('arial', 10, 'bold'), bd=0).grid(row=29, column=3)
        self.linha_cont = Entry(inner_frame, font=('arial', 10, 'bold'), bd=2, width=3)
        self.linha_cont.grid(row=30, column=3, pady=10)
        self.linha_cont.insert(0, 0)


    def tela_observacoes(self):
        self.observacoes = Toplevel()
        titulo = ' '
        self.observacoes.title(160 * titulo + 'Observações')
        self.observacoes.geometry('1100x680+200+20')
        self.observacoes.config(bg='DeepSkyBlue2')

        self.obs_frame = Frame(self.observacoes, width=1080, height=580, relief=RIDGE, bd=7)
        self.obs_frame.place(x=10, y=10)
        self.obs_frame1 = Frame(self.observacoes, width=1080, height=70, relief=RIDGE, bd=7)
        self.obs_frame1.place(x=10, y=590)

        Label(self.obs_frame, text='Obs. Materiais: ', font=self.fonte, bd=0).place(x=100, y=50)
        Label(self.obs_frame, text='Obs. Serviços: ', font=self.fonte, bd=0).place(x=100, y=180)
        Label(self.obs_frame, text='Obs. 1: ', font=self.fonte, bd=0).place(x=100, y=320)
        Label(self.obs_frame, text='Obs. 2: ', font=self.fonte, bd=0).place(x=100, y=470)


        self.obs = Text(self.obs_frame, width=75, height=6, bd=4, font='arial', wrap=WORD)
        self.obs.place(x=250, y=50)
        self.obs.insert(1.0, 'Produto para Consumo Final. \nFabricante: Alíquota de ICMS de 18% conforme RICMS-SP/2000, '
                             'Livro I, Título III, Capítulo II, Seção II, Artigo 52, Inciso I \nRevendedor: Informar o '
                             'ICMS-ST  recolhido anteriormente\n\nEntrega de materiais em local diverso do destinatário: '
                             'o endereço deverá constar na nota fiscal em campo específico do xml (bloco G) e em dados '
                             'adicionais. (Regime Especial 28558/2018).')

        self.obs_serv = Text(self.obs_frame, width=75, height=6, bd=4, font='arial', wrap=WORD)
        self.obs_serv.place(x=250, y=180)

        self.obs1 = Text(self.obs_frame, width=75, height=6, bd=4, font='arial', wrap=WORD)
        self.obs1.place(x=250, y=320)
        self.obs1.insert(1.0,
                         'Obs 1: Caso o fornecedor possua alguma especificidade que implique tratamento tributário diverso '
                         'do exposto acima, ou seja do regime tributário "SIMPLES NACIONAL" deverá  apresentar '
                         'documentação hábil que comprove sua condição peculiar, a qual será alvo de análise '
                         'prévia pela GECOT.')
        self.obs2 = Text(self.obs_frame, width=75, height=3, bd=4, font='arial', wrap=WORD)
        self.obs2.place(x=250, y=470)
        self.obs2.insert(1.0,
                         'Obs 2: Essa Análise não é exaustiva, podendo sofrer alterações no decorrer do processo de '
                         'contratação em relação ao produto/serviço.')

        self.chama_princ2 = Button(self.obs_frame1, font=('arial', 14, 'bold'), text='Dados Iniciais', bd=3, width=20,
                                  command=lambda: [self.principal.deiconify(), self.observacoes.withdraw()]).place(x=20,
                                                                                                                 y=2)

        self.chama_serv3 = Button(self.obs_frame1, font=('arial', 14, 'bold'), text='Serviços', bd=3, width=20,
                                  command=lambda: [self.servicos.deiconify(), self.observacoes.withdraw(),
                                                   self.principal.withdraw()]).place(x=270, y=2)

        self.chama_mat3 = Button(self.obs_frame1, font=('arial', 14, 'bold'), text='Materiais', bd=3, width=20,
                                command=lambda: [self.materiais.deiconify(), self.observacoes.withdraw(),
                                                 self.principal.withdraw()]).place(x=520, y=2)

        self.chama_contr3 = Button(self.obs_frame1, font=('arial', 14, 'bold'), text='Clausulas', bd=3, width=20,
                                   command=lambda: [self.tela_contratos(), self.observacoes.withdraw(),
                                                    self.principal.withdraw()]).place(x=770, y=2)




    def salvar(self):
        # ============================== GUARDAR DADOS ========================================#
        dados_salvos = []
        dados_salvos.extend([datetime.now().strftime('%d/%m/%Y, %H:%M:%S'), self.gere.get(), self.proc.get(),
                             self.req.get(), self.orcam.get(), self.objcust.get(), self.tipo1.get(), self.tipo2.get(),
                             self.tipo3.get(), self.objeto.get(1.0, END), self.valor.get(), self.complem.get(1.0, END),
                             self.data_mat, self.linha_mat.get(), self.serv.get(1.0, END), self.iva.get(),
                             self.data, self.linha_serv.get(), self.obs.get(1.0, END), self.obs_serv.get(1.0, END),
                             self.obs1.get(1.0, END), self.obs2.get(1.0, END), [i.get(1.0, END) for i in self.infos],
                             [i.get() for i in self.lista_check]])

        with open("G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Base.txt",
                  "rb+") as fp:  # Pickling
            pickle_list = []
            nova_lista = []
            while True:
                try:
                    pickle_list.append(pickle.load(fp))
                except EOFError:
                    break
            for i in pickle_list:
                if i[2] != dados_salvos[2]:
                    nova_lista.append(i)
        with open("G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Base.txt", "wb") as fp:
            nova_lista.append(dados_salvos)
            for lis in nova_lista:
                pickle.dump(lis, fp)





        # with open("G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Base.txt", "rb") as fp:  # Pickling
        #     pickle_list = []
        #     nova_lista = []
        #     while True:
        #         try:
        #             pickle_list.append(pickle.load(fp))
        #         except EOFError:
        #             break
            # for i in pickle_list:
            #     if len(pickle_list) > 1:
            #         if i[2] != dados_salvos[2]:
            #             nova_lista.append(i)
            #     else:
            #         pickle_list = pickle_list[0]
            #         for i in pickle_list:
            #             if i[2] != dados_salvos[2]:
            #                 nova_lista.append(i)
            #
            # nova_lista.append(dados_salvos)
            # pickle.dump(nova_lista, fp)


        # ============================== CRIAR PDF ============================================#
        self.pdf = FPDF(orientation='P', unit='mm', format='A4')
        self.pdf.add_page()

        self.pdf_w = 210
        self.pdf_h = 297
        self.pdf.rect(5.0, 5.0, 200.0, 280.0)
        self.pdf.rect(5.0, 5.0, 200.0, 20.0)

        self.pdf.image('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\logo.jpg', x=7.0, y=7.0,
                       h=15.0, w=50.0)
        self.pdf.line(70.0, 5.0, 70.0, 25.0)

        self.pdf.set_font('Arial', 'B', 10)
        self.pdf.set_xy(75.0, 9.0)
        self.pdf.multi_cell(w=125, h=5,
                            txt='Análise Contábil e Tributária para Processos de Licitação e ou Contratação Direta')

        self.pdf.rect(5.0, 30.0, 200.0, 25.0)
        self.pdf.line(5.0, 40.0, 205.0, 40.0)
        self.pdf.line(88.0, 30.0, 88.0, 55.0)
        self.pdf.line(137.0, 30.0, 137.0, 40.0)

        # ===================================== INFORMAÇÕES INICIAIS ==================================#
        self.pdf.set_xy(10.0, 25.0)
        self.pdf.cell(w=40, h=20, txt='Gerência Contratante:')
        self.pdf.set_xy(50.0, 25.0)
        self.pdf.cell(w=40, h=20, txt=self.gere.get())
        self.pdf.set_xy(90.0, 22.5)
        self.pdf.cell(w=40, h=20, txt='N° do Processo GECBS:')
        self.pdf.set_xy(90.0, 26.5)
        self.pdf.cell(w=40, h=20, txt=self.proc.get())
        self.pdf.set_xy(142.0, 22.5)
        self.pdf.cell(w=40, h=20, txt='Requisição de Compras: ')
        self.pdf.set_xy(142.0, 26.5)
        self.pdf.cell(w=40, h=20, txt=self.req.get())
        self.pdf.set_xy(10.0, 35.0)
        self.pdf.cell(w=40, h=20, txt='Objeto de Custos:')
        self.pdf.set_xy(43.0, 42.5)
        self.pdf.multi_cell(w=45, h=5, align='L', txt=self.objcust.get())
        self.pdf.set_xy(90.0, 35.0)
        self.pdf.cell(w=40, h=20, txt='Consta no Orçamento? ')
        self.pdf.set_xy(148.0, 35.0)
        self.pdf.cell(w=40, h=20, txt='Sim')
        self.pdf.set_xy(175.0, 35.0)
        self.pdf.cell(w=40, h=20, txt='Não')
        if self.orcam.get() == 'Sim':
            self.pdf.set_xy(138.0, 35.0)
            self.pdf.cell(w=40, h=20, txt='X')
        else:
            self.pdf.set_xy(169.0, 35.0)
            self.pdf.cell(w=40, h=20, txt='X')
        # self.pdf.rect(5.0, 60.0, 200.0, 225.0)
        self.pdf.line(5.0, 66.0, 205.0, 66.0)
        self.pdf.set_xy(100.0, 51.0)
        self.pdf.cell(w=40, h=20, txt='ANÁLISE')
        self.pdf.line(5.0, 75.0, 205.0, 75.0)
        self.pdf.set_xy(30.0, 60.5)
        self.pdf.cell(w=40, h=20, txt='Serviço')
        self.pdf.set_xy(90.0, 60.5)
        self.pdf.cell(w=40, h=20, txt='Material')
        self.pdf.set_xy(135.0, 60.5)
        self.pdf.cell(w=40, h=20, txt='Serviço com Fornecimento de Material')
        if self.tipo1.get() == 1:
            self.pdf.set_xy(25.0, 60.5)
            self.pdf.cell(w=40, h=20, txt='X')
        if self.tipo2.get() == 1:
            self.pdf.set_xy(85.0, 60.5)
            self.pdf.cell(w=40, h=20, txt='X')
        if self.tipo3.get() == 1:
            self.pdf.set_xy(130.0, 60.5)
            self.pdf.cell(w=40, h=20, txt='X')
        self.pdf.set_xy(10.0, 77.0)
        self.pdf.cell(w=40, h=5, txt='Objeto: ')
        self.pdf.set_xy(30.0, 77.0)
        self.pdf.set_font('')
        self.pdf.multi_cell(w=160, h=5, txt=self.objeto.get(1.0, 'end'))
        self.pdf.set_font('arial', 'B', 10)
        self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
        self.pdf.cell(w=40, h=5, txt='Valor estimado: ')
        self.pdf.set_xy(40.0, self.pdf.get_y())
        self.pdf.set_font('')
        self.pdf.cell(w=40, h=5, txt=self.valor.get())
        self.pdf.set_font('arial', 'B', 10)
        self.pdf.set_xy(15.0, self.pdf.get_y() + 10)
        self.pdf.cell(w=40, h=5, txt=self.complem.get(1.0, END))
        self.pdf.set_xy(15.0, self.pdf.get_y() + 10)

        self.pdf.set_auto_page_break(True, 20.0)
        # self.pdf.set_auto_page_break(True, 20.0)

        # ======================================= MATERIAIS =================================#

        if len(self.data_mat) > 1:
            self.pdf.set_xy(10, self.pdf.get_y() + 5)
            if self.pdf.get_y() > 276:
                self.pdf.rect(5.0, 5.0, 200.0, 280.0)


            q1 = self.pdf.get_y()
            self.pdf.set_font('')
            self.pdf.multi_cell(w=180, h=5, txt='Informações Tributárias: ')
            self.pdf.set_xy(10, self.pdf.get_y() + 5)
            self.pdf.multi_cell(w=180, h=5, txt=' - Materiais serão acobertados por Nota Fiscal eletrônica modelo 55. ')
            self.pdf.set_font('Arial', 'B', 10)

            self.dados_faltantes = []
            # def cria_tabela(data):
            self.pdf.set_xy(10, self.pdf.get_y() + 5)
            cont_lista = 0
            cont = 1
            px = 10
            py = self.pdf.get_y()
            # while self.pdf.get_y() < 274:
            for row in self.data_mat:
                for datum in row:
                    if cont == 1:
                        self.pdf.set_font('') if cont_lista != 0 else self.pdf.set_font('Arial', 'B', 10)
                        self.pdf.set_xy(px, py)
                        self.pdf.multi_cell(w=20, h=5, txt=datum, border=1)
                    elif cont == 2:
                        self.pdf.set_xy(px + 20, py)
                        self.pdf.multi_cell(w=75, h=5, txt=datum, border=1)
                    elif cont == 3:
                        self.pdf.set_xy(px + 95, py)
                        self.pdf.multi_cell(w=10, h=5, txt=datum, border=1)
                    elif cont == 4:
                        self.pdf.set_xy(px + 105, py)
                        self.pdf.multi_cell(w=20, h=5, txt=datum, border=1)
                    else:
                        self.pdf.set_xy(px + 125, py)
                        px += 16
                        self.pdf.multi_cell(w=16, h=5, txt=datum, border=1)
                    cont += 1
                px = 10
                py = self.pdf.get_y()
                cont = 1
                cont_lista += 1
                if py > 270:
                    self.dados_faltantes = self.data_mat[cont_lista:]
                    self.pdf.add_page()
                    self.pdf.rect(5.0, 5.0, 200.0, 280.0)
                    break

            if self.dados_faltantes:
                self.dados_faltantes.insert(0, ['DESCRIÇÃO', 'CÓDIGO', 'C.C'])
                cont = 3
                px = 10
                py = self.pdf.get_y()
                for row in self.dados_faltantes:
                    for datum in row:
                        if cont % 3 == 0:
                            self.pdf.set_xy(px + 20, py)
                            self.pdf.multi_cell(w=150, h=5, txt=datum, border=1)
                        elif cont % 4 == 0:
                            atual = self.pdf.get_y() - py
                            self.pdf.set_xy(px, py)
                            self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                        else:
                            atual = self.pdf.get_y() - py
                            self.pdf.set_xy(px + 170, py)
                            self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                        cont += 1
                    px = 10
                    py = self.pdf.get_y()
                    cont = 3

            else:
                pass

            self.pdf.rect(7.5, q1 - 3, 195, self.pdf.get_y() - q1 + 7.5)
            self.pdf.set_xy(10.0, self.pdf.get_y() + 10)
            self.pdf.set_font('Arial', 'U', 10)
            self.pdf.multi_cell(w=180, h=5, txt=self.obs.get(1.0, 'end'))
            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
            self.pdf.set_font('')
            self.pdf.rect(5.0, 5.0, 200.0, 280.0)
        # ======================================== SERVIÇOS =============================================#




        if self.iva.get() != '':


            self.pdf.set_y(self.pdf.get_y() + (float(self.linha_serv.get()) * 10))
            if int(self.linha_serv.get()) > 0:
                self.pdf.add_page()
                self.pdf.rect(5.0, 5.0, 200.0, 280.0)

            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
            q2 = self.pdf.get_y()
            print(q2)
            self.pdf.multi_cell(w=180, h=5,
                                txt=' - Serviços serão acobertados por Nota Fiscal de serviço eletrônica. ')

            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
            self.pdf.multi_cell(w=180, h=5, txt='O código de imposto (IVA) utilizado no pedido (SAP) '
                                            'deverá ser o ' + self.iva.get() + '.')




            self.dados_faltantes = []
            # def cria_tabela(data):
            self.pdf.set_xy(10, self.pdf.get_y() + 5)
            cont_lista = 0
            cont = 3
            px = 10
            py = self.pdf.get_y()
            # while self.pdf.get_y() < 274:
            for row in self.data:
                for datum in row:
                    if cont % 3 == 0:
                        self.pdf.set_font('') if cont_lista != 0 else self.pdf.set_font('Arial', 'B', 10)
                        self.pdf.set_xy(px + 20, py)
                        self.pdf.multi_cell(w=150, h=5, txt=datum, border=1)
                    elif cont % 4 == 0:
                        atual = self.pdf.get_y() - py
                        self.pdf.set_xy(px, py)
                        self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                    else:
                        atual = self.pdf.get_y() - py
                        self.pdf.set_xy(px + 170, py)
                        self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                    cont += 1
                px = 10
                py = self.pdf.get_y()
                cont = 3
                cont_lista += 1
                if py > 270:
                    self.dados_faltantes = self.data[cont_lista:]
                    self.pdf.rect(7.5, q2 - 3, 195.0, self.pdf.get_y() - q2 + 7.5)
                    self.pdf.add_page()
                    self.pdf.rect(5.0, 5.0, 200.0, 280.0)
                    break

            if self.dados_faltantes:
                q2 = self.pdf.get_y()
                self.dados_faltantes.insert(0, ['DESCRIÇÃO', 'CÓDIGO', 'C.C'])
                cont = 3
                px = 10
                py = self.pdf.get_y()
                # while self.pdf.get_y() < 274:
                for row in self.dados_faltantes:
                    for datum in row:
                        if cont % 3 == 0:
                            self.pdf.set_xy(px + 20, py)
                            self.pdf.multi_cell(w=150, h=5, txt=datum, border=1)
                        elif cont % 4 == 0:
                            atual = self.pdf.get_y() - py
                            self.pdf.set_xy(px, py)
                            self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                        else:
                            atual = self.pdf.get_y() - py
                            self.pdf.set_xy(px + 170, py)
                            self.pdf.multi_cell(w=20, h=atual, txt=datum, border=1)
                        cont += 1
                    px = 10
                    py = self.pdf.get_y()
                    cont = 3

            else:
                pass




            self.pdf.set_xy(15.0, self.pdf.get_y() + 10)

            self.pdf.multi_cell(w=180, h=5, txt='Informações Tributárias: ')

            # self.pdf.set_font('Arial', 'B', 10)
            self.pdf.set_xy(15.0, self.pdf.get_y() + 5)
            self.pdf.set_font('Arial', 'B', 10)
            self.pdf.multi_cell(w=180, h=5, txt=self.serv.get(1.0, 'end'))
            self.pdf.rect(7.5, q2 - 3, 195.0, self.pdf.get_y() - q2 + 7.5)
            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
            self.pdf.multi_cell(w=180, h=5, txt=self.obs_serv.get(1.0, 'end'))
            self.pdf.set_xy(10.0, self.pdf.get_y() + 5) if self.obs_serv.get(1.0, 'end') != '' else \
                self.pdf.set_xy(10.0, self.pdf.get_y())
            self.pdf.rect(5.0, 5.0, 200.0, 280.0)


        # ============================== OBSERVAÇÕES ===========================================#
        # print(self.pdf.get_y())
        # if self.pdf.get_y() > 270:
        #     self.pdf.add_page()
        #     self.pdf.rect(5.0, 5.0, 200.0, 280.0)
        # else:
        #     pass
        self.pdf.rect(5.0, 5.0, 200.0, 280.0)
        self.pdf.set_font('')
        self.pdf.multi_cell(w=180, h=5, txt=self.obs1.get(1.0, 'end'))
        self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
        self.pdf.multi_cell(w=180, h=5, txt=self.obs2.get(1.0, 'end'))

        #=============================  INFORMAÇÕES CONTRATUAIS ===========================================#
        # print(self.pdf.get_y())
        # if self.pdf.get_y() > 270:
        #     self.pdf.add_page()
        #     self.pdf.rect(5.0, 5.0, 200.0, 285.0)
        # else:
        #     pass
        self.pdf.set_y(self.pdf.get_y() + (float(self.linha_cont.get())*10))
        total = 0
        for l in self.lista_check:
            total += l.get()
        if total > 0:
            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
            self.pdf.cell(w=40, h=5, txt='Informações Contratuais : ')
            self.pdf.set_xy(10.0, self.pdf.get_y() + 5)
        for i, item in enumerate(self.lista_check):
            if item.get() == 1:
                self.pdf.set_xy(15.0, self.pdf.get_y() + 5)
                self.pdf.multi_cell(w=180, h=5, align='L', txt=self.infos[i].get(1.0, 'end'))
                if self.pdf.get_y() > 270:
                    self.pdf.add_page()
                    self.pdf.rect(5.0, 5.0, 200.0, 285.0)

        if int(self.linha_cont.get()) > 0:
            self.pdf.rect(5.0, 5.0, 200.0, 280.0)

        # if self.pdf.get_y() > 265:
        #     self.pdf.add_page()
        #     self.pdf.rect(5.0, 5.0, 200.0, 280.0)



        # =============================== ASSINATURA E DATA =====================================#
        self.pdf.rect(5.0, 265.0, 200.0, 20.0)
        self.pdf.set_xy(10.0, 270.0)
        self.pdf.cell(w=40, txt='Responsável pela análise: ')
        self.pdf.line(80.0, 265.0, 80.0, 285.0)
        self.pdf.set_xy(90.0, 270.0)
        self.pdf.cell(w=40, txt='DATA: ' + date.today().strftime('%d/%m/%Y'))
        self.pdf.line(130.0, 265.0, 130.0, 285.0)
        self.pdf.set_xy(135.0, 270.0)
        self.pdf.cell(w=40, txt='Revisado pela Gerência: ')
        self.pdf.image('G:\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\Mari.png', x=7.0, y=265.0, h=25.0, w=45.0)
        troca = self.proc.get().replace('/', '-')
        path = '\\\GBD_VT1NTAQA\Data2\GECOT\Análise Contábil_Tributária_Licitações\\2021\\1Pendentes\\'
        self.pdf.output(path + 'Análise Tributária - ' + troca + '.pdf', 'F')
        os.startfile(path + 'Análise Tributária - ' + troca + '.pdf')



if __name__=='__main__':
    janela = Tk()
    aplicacao = Analise(janela)
    janela.mainloop()

